#!/usr/bin/python3
# -*- encoding: utf-8 -*-

import openpyxl
from openpyxl.styles import Font

# ワークブックを新規作成
wb = openpyxl.Workbook()
print(str(wb.get_sheet_names()))
sheet = wb.active
print(str(sheet.title))
sheet.title = 'Spam Bacon Eggs Sheet'
print(str(wb.get_sheet_names()))
wb.save('saving_sample.xlsx')

# シートを生成
wb.create_sheet(index = 0, title = 'First Sheet')
wb.create_sheet(index = 2, title = 'Middle Sheet')
wb.create_sheet(index = 3, title = 'Last Sheet')
print(str(wb.get_sheet_names()))
wb.save('saving_sample.xlsx')

# シートを削除
wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))
wb.save('saving_sample.xlsx')

# セルに値を記入
sheet = wb.get_sheet_by_name('First Sheet')
sheet['A1'] = 'Hello world!'

# フォントを変更
italic24Font = Font(size = 24, italic = True)
sheet['B2'].font = italic24Font
sheet['B2'] = 'Hello world!'
wb.save('saving_sample.xlsx')

