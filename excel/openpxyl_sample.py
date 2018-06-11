#!/usr/bin/python3
# -*- encoding: utf-8 -*-

import openpyxl
#from openpyxl.cell import get_column_letter, column_index_from_string


# Workbookを取得
wb = openpyxl.load_workbook('sample.xlsx')
print('type of wb is ' + str(type(wb)))

# Worksheetを取得
hogeSheet = wb.get_sheet_by_name('hogesheet')
print('type of sheet is ' + str(type(hogeSheet)))
print('hogeSheet title is ' + hogeSheet.title)
activeSheet = wb.active
print('active sheet title is ' + activeSheet.title)

# Cell内容を取得。valueのデータ型は色々
print('type of cell is ' + str(type(hogeSheet['B2'])))
print('B2 is ' + str(hogeSheet['B2'].value) + ', type is ' + str(type(hogeSheet['B2'].value)))
print('B8 is ' + str(hogeSheet['B8'].value) + ', type is ' + str(type(hogeSheet['B8'].value)))
print('B9 is ' + str(hogeSheet['B9'].value) + ', type is ' + str(type(hogeSheet['B9'].value)))

# Cellからインデックス情報を取得
cell = hogeSheet['B2']
print('Row index is ' + str(cell.row) + ', Column index is ' + str(cell.column))

# C4セルの内容を出力する インデックス指定は1オリジン
print('cell(4,3) is ' + str(hogeSheet.cell(row = 4, column = 3).value))
for i in range(3, 6, 1):
    print(i, hogeSheet.cell(row = i, column = 3).value)

# データ入力済セルの範囲を取得
print('max_row = ' + str(hogeSheet.max_row) + ', max_column = ' + str(hogeSheet.max_column))

# 列番号をアルファベットに変換する TODO import 失敗
#print(get_column_letter(1))

# シートを範囲指定し、loopでcellを回す
for row in tuple(hogeSheet['b2':'c6']):
    for cell in row:
        print(cell.coordinate, cell.value)
    print('END OF ROW')

# シートを列選択 TODO TypeError: 'generator' object is not subscriptable
hogeSheet.columns[1]
#for celll in hogeSheet.columns[2]:
#    print(str(celll.value))

